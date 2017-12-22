#!/usr/bin/python

from openpyxl import load_workbook
import json


import os
os.path.join('usr','bin','dir')


for filename in os.listdir('./json'):
    #print(filename)
    file = os.path.join('./json', filename)
    #print(file)

    f = open(file,"r")


    js=f.read().replace('\n', '')
    f.close()
    os.remove(file)
    #print(js)
    ds = json.loads(js)


    wb = load_workbook("bill.xlsx")
    ws = wb.get_sheet_by_name("s1")
    c = ws.cell(row = 10, column = 5)
    c.value = ds["bill-no"];
    c = ws.cell(row = 10, column = 1)
    c.value = ds["billto"];
    c = ws.cell(row = 10, column = 6)
    c.value = ds["Date"];
    #print(ds["details"])
    ct=14
    for i in ds["details"]:
        c = ws.cell(row = ct, column = 1)
        c.value = i["desc"];
        c = ws.cell(row = ct, column = 3)
        c.value = i["qty"];
        c = ws.cell(row = ct, column = 5)
        c.value = i["price"];
        ct+=1;

    wb.save("./bills/"+str(ds["bill-no"])+".xlsx")